"""
PaperPilot - Excel Exporter
Phase 6.2: Export records, screening log, extraction data, and summary to Excel.
"""

from __future__ import annotations

import subprocess
import base64
import json
from datetime import datetime
from typing import Any

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None  # type: ignore


# ---------------------------------------------------------------------------
# Color constants
# ---------------------------------------------------------------------------
HEADER_BG = "1F3864"   # deep blue
HEADER_FG = "FFFFFF"   # white
INCLUDE_BG = "C6EFCE"  # light green
EXCLUDE_BG = "FFC7CE"  # light red
MAYBE_BG   = "FFEB9C"  # light yellow (optional, for completeness)


def _header_font() -> "Font":
    return Font(bold=True, color=HEADER_FG, name="Calibri", size=11)


def _header_fill() -> "PatternFill":
    return PatternFill("solid", fgColor=HEADER_BG)


def _row_fill(color: str) -> "PatternFill":
    return PatternFill("solid", fgColor=color)


def _center() -> "Alignment":
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _left() -> "Alignment":
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def _auto_width(ws, min_w: int = 8, max_w: int = 50) -> None:
    """Adjust column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                try:
                    cell_len = len(str(cell.value))
                except Exception:
                    cell_len = 0
                if cell_len > max_len:
                    max_len = cell_len
        adjusted = max(min_w, min(max_len + 2, max_w))
        ws.column_dimensions[col_letter].width = adjusted


def _write_header(ws, headers: list[str]) -> None:
    """Write styled header row."""
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _header_font()
        cell.fill = _header_fill()
        cell.alignment = _center()


def _str(value: Any, default: str = "") -> str:
    """Safe string conversion."""
    if value is None:
        return default
    if isinstance(value, list):
        return "; ".join(str(v) for v in value)
    return str(value)


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def _build_records_sheet(ws, records: list[dict],
                          decisions: dict[str, dict]) -> int:
    """Sheet 1: Records — returns row count (excluding header)."""
    headers = [
        "#", "Title", "Authors", "Year", "Journal",
        "DOI", "PMID", "Keywords", "Decision", "Reason", "Score", "Has_PDF"
    ]
    _write_header(ws, headers)
    ws.freeze_panes = "A2"

    row_count = 0
    for idx, rec in enumerate(records, start=1):
        rec_id = _str(rec.get("id") or rec.get("record_id") or rec.get("pmid") or idx)
        decision_rec = decisions.get(rec_id, {})
        decision_val = _str(decision_rec.get("decision", "undecided")).lower()
        reason_val   = _str(decision_rec.get("reason", ""))
        score_val    = decision_rec.get("score", "")
        has_pdf      = "Yes" if rec.get("has_pdf") or rec.get("pdf_path") else "No"

        # Authors: support list or string
        authors_raw = rec.get("authors", rec.get("author", ""))
        if isinstance(authors_raw, list):
            authors_str = "; ".join(
                (a.get("name", "") if isinstance(a, dict) else _str(a))
                for a in authors_raw
            )
        else:
            authors_str = _str(authors_raw)

        keywords_raw = rec.get("keywords", rec.get("keyword", ""))
        if isinstance(keywords_raw, list):
            keywords_str = "; ".join(
                (k.get("keyword", k) if isinstance(k, dict) else _str(k))
                for k in keywords_raw
            )
        else:
            keywords_str = _str(keywords_raw)

        row = [
            idx,
            _str(rec.get("title", "")),
            authors_str,
            _str(rec.get("year", rec.get("pub_year", ""))),
            _str(rec.get("journal", rec.get("source", rec.get("journal_name", "")))),
            _str(rec.get("doi", "")),
            _str(rec.get("pmid", rec.get("pubmed_id", ""))),
            keywords_str,
            decision_val,
            reason_val,
            score_val,
            has_pdf,
        ]
        ws.append(row)
        row_count += 1

        # Apply row styling
        data_row = ws[ws.max_row]
        fill = None
        if decision_val == "include":
            fill = _row_fill(INCLUDE_BG)
        elif decision_val == "exclude":
            fill = _row_fill(EXCLUDE_BG)
        elif decision_val == "maybe":
            fill = _row_fill(MAYBE_BG)

        for cell in data_row:
            cell.alignment = _left()
            if fill:
                cell.fill = fill

    _auto_width(ws)
    return row_count


def _build_screening_log_sheet(ws, records: list[dict],
                                decision_history: dict[str, list[dict]]) -> int:
    """Sheet 2: Screening Log — returns row count."""
    headers = [
        "Record_ID", "Title", "Stage", "Decision", "Reason_Code", "Timestamp"
    ]
    _write_header(ws, headers)

    # Build title lookup
    title_map: dict[str, str] = {}
    for rec in records:
        rec_id = _str(
            rec.get("id") or rec.get("record_id") or rec.get("pmid") or ""
        )
        if rec_id:
            title_map[rec_id] = _str(rec.get("title", ""))

    row_count = 0
    for rec_id, history in decision_history.items():
        title = title_map.get(rec_id, "")
        if not isinstance(history, list):
            continue
        for entry in history:
            if not isinstance(entry, dict):
                continue
            row = [
                rec_id,
                title,
                _str(entry.get("stage", entry.get("screening_stage", ""))),
                _str(entry.get("decision", "")),
                _str(entry.get("reason_code", entry.get("reason", ""))),
                _str(entry.get("timestamp", entry.get("created_at", ""))),
            ]
            ws.append(row)
            row_count += 1
            for cell in ws[ws.max_row]:
                cell.alignment = _left()

    _auto_width(ws)
    return row_count


def _build_extraction_sheet(ws, records: list[dict],
                              extracted_values: dict[str, list[dict]]) -> int:
    """Sheet 3: Extraction — returns row count."""
    headers = [
        "Record_ID", "Title", "Field_Name", "Value", "Source", "Confidence"
    ]
    _write_header(ws, headers)

    title_map: dict[str, str] = {}
    for rec in records:
        rec_id = _str(
            rec.get("id") or rec.get("record_id") or rec.get("pmid") or ""
        )
        if rec_id:
            title_map[rec_id] = _str(rec.get("title", ""))

    row_count = 0
    for rec_id, extractions in extracted_values.items():
        title = title_map.get(rec_id, "")
        if not isinstance(extractions, list):
            continue
        for item in extractions:
            if not isinstance(item, dict):
                continue
            row = [
                rec_id,
                title,
                _str(item.get("field_name", item.get("field", ""))),
                _str(item.get("value", "")),
                _str(item.get("source", "")),
                _str(item.get("confidence", "")),
            ]
            ws.append(row)
            row_count += 1
            for cell in ws[ws.max_row]:
                cell.alignment = _left()

    _auto_width(ws)
    return row_count


def _build_summary_sheet(ws, records: list[dict],
                          decisions: dict[str, dict]) -> None:
    """Sheet 4: Summary — project statistics."""
    # Count decisions
    counts: dict[str, int] = {
        "include": 0, "exclude": 0, "maybe": 0, "undecided": 0
    }
    for rec in records:
        rec_id = _str(
            rec.get("id") or rec.get("record_id") or rec.get("pmid") or ""
        )
        dec = decisions.get(rec_id, {})
        val = _str(dec.get("decision", "undecided")).lower()
        if val in counts:
            counts[val] += 1
        else:
            counts["undecided"] += 1

    total = len(records)
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Write summary rows
    header_font = Font(bold=True, name="Calibri", size=11)
    header_fill = PatternFill("solid", fgColor="D9E1F2")  # light blue-grey

    rows = [
        ("Metric", "Value"),
        ("Total Records", total),
        ("Include",   counts["include"]),
        ("Exclude",   counts["exclude"]),
        ("Maybe",     counts["maybe"]),
        ("Undecided", counts["undecided"]),
        ("",          ""),
        ("Generated At", generated_at),
        ("Exporter", "PaperPilot Excel Exporter v1.0"),
    ]

    for i, (k, v) in enumerate(rows, start=1):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)

        cell_k = ws.cell(row=i, column=1)
        cell_v = ws.cell(row=i, column=2)

        if i == 1:
            # Header row
            cell_k.font = _header_font()
            cell_k.fill = _header_fill()
            cell_v.font = _header_font()
            cell_v.fill = _header_fill()
        elif k:
            cell_k.font = header_font
            cell_k.fill = header_fill

        cell_k.alignment = _left()
        cell_v.alignment = _left()

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30


# ---------------------------------------------------------------------------
# Main export function
# ---------------------------------------------------------------------------

def export_excel(
    records: list[dict],
    decisions: dict[str, dict],
    decision_history: dict[str, list[dict]],
    extracted_values: dict[str, list[dict]],
    out_path: str,
) -> dict:
    """
    Export PaperPilot data to a multi-sheet Excel workbook.

    Parameters
    ----------
    records          : All bibliographic records.
    decisions        : Mapping record_id -> latest decision dict.
    decision_history : Mapping record_id -> list of historical decision dicts.
    extracted_values : Mapping record_id -> list of extracted field dicts.
    out_path         : Destination file path (*.xlsx).

    Returns
    -------
    dict with keys ``sheets`` (int) and ``rows`` (int, total data rows written).
    """
    if openpyxl is None:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install it with: pip install openpyxl"
        )

    wb = Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Sheet 1: Records
    ws1 = wb.create_sheet("Records")
    rows1 = _build_records_sheet(ws1, records, decisions)

    # Sheet 2: Screening Log
    ws2 = wb.create_sheet("Screening Log")
    rows2 = _build_screening_log_sheet(ws2, records, decision_history)

    # Sheet 3: Extraction (only if data present)
    ws3 = wb.create_sheet("Extraction")
    rows3 = _build_extraction_sheet(ws3, records, extracted_values)

    # Sheet 4: Summary
    ws4 = wb.create_sheet("Summary")
    _build_summary_sheet(ws4, records, decisions)

    wb.save(out_path)

    total_rows = rows1 + rows2 + rows3
    return {"sheets": 4, "rows": total_rows}
